import polars as pl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# Define type mappings for equivalent data types
TYPE_MAPPINGS = {
    'INTEGER': ['INT', 'INTEGER', 'BIGINT', 'SMALLINT', 'TINYINT'],
    'VARCHAR': ['VARCHAR', 'TEXT', 'CHAR', 'STRING', 'NVARCHAR'],
    'DECIMAL': ['DECIMAL', 'NUMERIC', 'NUMBER'],
    'FLOAT': ['FLOAT', 'REAL', 'DOUBLE', 'DOUBLE PRECISION'],
    'TIMESTAMP': ['TIMESTAMP', 'DATETIME', 'DATETIME2'],
    'DATE': ['DATE'],
    'BOOLEAN': ['BOOLEAN', 'BOOL', 'BIT']
}

def are_types_compatible(type1, type2):
    """Check if two data types are considered compatible"""
    type1, type2 = type1.upper(), type2.upper()
    
    # Strip length specifications like VARCHAR(50) to VARCHAR
    type1 = type1.split('(')[0].strip()
    type2 = type2.split('(')[0].strip()
    
    # If types are exactly the same, they're compatible
    if type1 == type2:
        return True
    
    # Check if types belong to the same group
    for type_group in TYPE_MAPPINGS.values():
        if type1 in type_group and type2 in type_group:
            return True
    
    return False

def read_csv_files():
    """Read source and target CSV files using Polars"""
    source_df = pl.read_csv("csv/source.csv")
    target_df = pl.read_csv("csv/target.csv")
    return source_df, target_df

def compare_tables(source_df, target_df):
    """Compare tables between source and target"""
    source_tables = set(source_df['name'].unique())
    target_tables = set(target_df['NAME'].unique())
    
    common_tables = source_tables.intersection(target_tables)
    source_only = source_tables - target_tables
    target_only = target_tables - source_tables
    
    return {
        'common': sorted(list(common_tables)),
        'source_only': sorted(list(source_only)),
        'target_only': sorted(list(target_only))
    }

def compare_columns(source_df, target_df, table_name):
    """Compare columns for a specific table"""
    source_cols = source_df.filter(pl.col('name') == table_name).select(['col_name', 'data_type', 'ordinal_position'])
    target_cols = target_df.filter(pl.col('NAME') == table_name).select(['COL_NAME', 'DATA_TYPE', 'ORDINAL_POSITION'])
    
    # Rename target columns to match source
    target_cols = target_cols.rename({
        'COL_NAME': 'col_name',
        'DATA_TYPE': 'data_type',
        'ORDINAL_POSITION': 'ordinal_position'
    })
    
    source_cols_set = set(source_cols['col_name'].to_list())
    target_cols_set = set(target_cols['col_name'].to_list())
    
    common_cols = source_cols_set.intersection(target_cols_set)
    source_only = source_cols_set - target_cols_set
    target_only = target_cols_set - source_cols_set
    
    # Compare data types for common columns
    datatype_mismatches = []
    for col in common_cols:
        source_type = source_cols.filter(pl.col('col_name') == col)['data_type'].item()
        target_type = target_cols.filter(pl.col('col_name') == col)['data_type'].item()
        if not are_types_compatible(source_type, target_type):
            datatype_mismatches.append({
                'column': col,
                'source_type': source_type,
                'target_type': target_type
            })
    
    return {
        'common': sorted(list(common_cols)),
        'source_only': sorted(list(source_only)),
        'target_only': sorted(list(target_only)),
        'datatype_mismatches': datatype_mismatches
    }

def format_worksheet(ws):
    """Apply formatting to worksheet"""
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Calculate width: add 2 for padding, cap at 21.6 (3 inches)
        # If content is smaller, use the smaller width
        adjusted_width = min(max_length + 2, 21.6)
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def create_excel_report(comparison_results, source_df, target_df):
    """Create formatted Excel report"""
    wb = Workbook()
    
    # Table Comparison Sheet
    ws_tables = wb.active
    ws_tables.title = "Table Comparison"
    ws_tables.append(['Category', 'Table Name'])
    
    for table in comparison_results['tables']['common']:
        ws_tables.append(['Common', table])
    for table in comparison_results['tables']['source_only']:
        ws_tables.append(['Source Only', table])
    for table in comparison_results['tables']['target_only']:
        ws_tables.append(['Target Only', table])
    
    format_worksheet(ws_tables)
    
    # Column Comparison Sheet
    ws_columns = wb.create_sheet("Column Comparison")
    ws_columns.append(['Table Name', 'Column Name', 'Status', 'Source Type', 'Target Type'])
    
    for table in comparison_results['columns']:
        # Get all columns from source and target for this table
        source_cols = source_df.filter(pl.col('name') == table['table_name']).select(['col_name', 'data_type'])
        target_cols = target_df.filter(pl.col('NAME') == table['table_name']).select(['COL_NAME', 'DATA_TYPE'])
        
        # Create dictionaries for easy lookup
        source_types = dict(zip(source_cols['col_name'], source_cols['data_type']))
        target_types = dict(zip(target_cols['COL_NAME'], target_cols['DATA_TYPE']))
        
        # Process all columns
        all_columns = sorted(set(list(source_types.keys()) + list(target_types.keys())))
        
        for col in all_columns:
            source_type = source_types.get(col, 'N/A')
            target_type = target_types.get(col, 'N/A')
            
            if col in table['source_only']:
                status = 'Source Only'
            elif col in table['target_only']:
                status = 'Target Only'
            else:  # Column exists in both
                if are_types_compatible(source_type, target_type):
                    status = 'Matching'
                else:
                    status = 'Different Types'
            
            ws_columns.append([
                table['table_name'],
                col,
                status,
                source_type,
                target_type
            ])
    
    format_worksheet(ws_columns)
    
    # Datatype Mismatches Sheet
    ws_datatypes = wb.create_sheet("Datatype Mismatches")
    ws_datatypes.append(['Table Name', 'Column Name', 'Source Type', 'Target Type'])
    
    for table in comparison_results['columns']:
        for mismatch in table['datatype_mismatches']:
            ws_datatypes.append([
                table['table_name'],
                mismatch['column'],
                mismatch['source_type'],
                mismatch['target_type']
            ])
    
    format_worksheet(ws_datatypes)
    
    # Save the workbook with timestamp
    os.makedirs("results", exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    wb.save(f"results/results_{timestamp}.xlsx")

def main():
    # Read source and target files
    source_df, target_df = read_csv_files()
    
    # Compare tables
    table_comparison = compare_tables(source_df, target_df)
    
    # Compare columns for common tables
    column_comparisons = []
    for table in table_comparison['common']:
        column_comparison = compare_columns(source_df, target_df, table)
        column_comparisons.append({
            'table_name': table,
            **column_comparison
        })
    
    # Create comparison results dictionary
    comparison_results = {
        'tables': table_comparison,
        'columns': column_comparisons
    }
    
    # Generate Excel report
    create_excel_report(comparison_results, source_df, target_df)

if __name__ == "__main__":
    main()

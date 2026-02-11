"""Column comparison tool — compare schemas between files or databases."""

import argparse
import re
import os
import logging
from datetime import datetime
from pathlib import Path

import polars as pl
import pyarrow
import pyarrow.parquet as pq
import yaml
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from dbqt.tools.utils import (
    load_config,
    setup_logging,
    Timer,
    fetch_metadata_parallel,
    _read_table_lists,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Type-mapping helpers
# ---------------------------------------------------------------------------

DEFAULT_TYPE_MAPPINGS = {
    "INTEGER": ["INT", "INTEGER", "BIGINT", "SMALLINT", "TINYINT", "NUMBER"],
    "VARCHAR": ["VARCHAR", "TEXT", "CHAR", "STRING", "NVARCHAR", "VARCHAR2"],
    "DECIMAL": ["DECIMAL", "NUMERIC", "NUMBER"],
    "FLOAT": ["FLOAT", "REAL", "DOUBLE", "DOUBLE PRECISION"],
    "TIMESTAMP": ["TIMESTAMP", "DATETIME"],
    "DATE": ["DATE", "TIMESTAMP"],
    "BOOLEAN": ["BOOLEAN", "BOOL", "BIT"],
}


def load_type_mappings(config_path=None):
    """Load type mappings from YAML config file or return defaults."""
    if config_path and Path(config_path).exists():
        logger.info(f"Loading type mappings from {config_path}")
        with open(config_path, "r") as f:
            config = yaml.safe_load(f)
            return config.get("type_mappings", DEFAULT_TYPE_MAPPINGS)
    return DEFAULT_TYPE_MAPPINGS


def generate_config_file(output_path="colcompare_config.yaml"):
    """Generate a default configuration file with type mappings."""
    config = {
        "type_mappings": DEFAULT_TYPE_MAPPINGS,
        "description": "Column comparison type mappings configuration. "
        "Each key represents a type group, and the list contains equivalent types.",
    }

    output_file = Path(output_path)
    if output_file.exists():
        logger.warning(f"Config file already exists at {output_path}")
        response = input("Overwrite? (y/n): ")
        if response.lower() != "y":
            logger.info("Config generation cancelled")
            return

    with open(output_path, "w") as f:
        yaml.dump(config, f, default_flow_style=False, sort_keys=False)

    logger.info(f"Generated default config file at {output_path}")
    print(f"\nDefault configuration saved to: {output_path}")
    print("You can now edit this file to customize type mappings.")


def are_types_compatible(type1, type2, type_mappings=None):
    """Check if two data types are considered compatible."""
    if type_mappings is None:
        type_mappings = DEFAULT_TYPE_MAPPINGS

    type1, type2 = type1.upper(), type2.upper()
    type1 = type1.split("(")[0].strip()
    type2 = type2.split("(")[0].strip()

    if type1 == type2:
        return True

    if re.match(r"^TIMESTAMP.*", type1) and re.match(r"^TIMESTAMP.*", type2):
        return True

    for type_group in type_mappings.values():
        if type1 in type_group and type2 in type_group:
            return True

    return False


# ---------------------------------------------------------------------------
# Parquet / file-based schema helpers
# ---------------------------------------------------------------------------


def _process_nested_type(field_type, parent_name="", processed_fields=None):
    """Recursively process nested types in Parquet schema."""
    if processed_fields is None:
        processed_fields = []

    if isinstance(field_type, (pyarrow.lib.ListType, pyarrow.lib.LargeListType)):
        element_type = field_type.value_type
        if isinstance(element_type, pyarrow.lib.StructType):
            for nested_field in element_type:
                full_name = (
                    f"{parent_name}__{nested_field.name}" if parent_name else nested_field.name
                )
                if isinstance(
                    nested_field.type,
                    (
                        pyarrow.lib.ListType,
                        pyarrow.lib.LargeListType,
                        pyarrow.lib.StructType,
                        pyarrow.lib.MapType,
                    ),
                ):
                    _process_nested_type(nested_field.type, full_name, processed_fields)
                else:
                    processed_fields.append({"col_name": full_name, "type": str(nested_field.type)})
        else:
            processed_fields.append({"col_name": parent_name, "type": str(field_type)})

    elif isinstance(field_type, pyarrow.lib.StructType):
        for nested_field in field_type:
            full_name = (
                f"{parent_name}__{nested_field.name}" if parent_name else nested_field.name
            )
            if isinstance(
                nested_field.type,
                (
                    pyarrow.lib.ListType,
                    pyarrow.lib.LargeListType,
                    pyarrow.lib.StructType,
                    pyarrow.lib.MapType,
                ),
            ):
                _process_nested_type(nested_field.type, full_name, processed_fields)
            else:
                processed_fields.append({"col_name": full_name, "type": str(nested_field.type)})

    elif isinstance(field_type, pyarrow.lib.MapType):
        processed_fields.append({"col_name": parent_name, "type": str(field_type)})

    return processed_fields


def _schema_to_df(schema, label="pq"):
    """Convert a pyarrow schema to a Polars DataFrame with SCH_TABLE/COL_NAME/DATA_TYPE."""
    fields = []
    for field in schema:
        if isinstance(
            field.type,
            (
                pyarrow.lib.ListType,
                pyarrow.lib.LargeListType,
                pyarrow.lib.StructType,
                pyarrow.lib.MapType,
            ),
        ):
            fields.extend(_process_nested_type(field.type, field.name))
        else:
            fields.append({"col_name": field.name, "type": str(field.type)})

    return pl.DataFrame(
        {
            "SCH_TABLE": [label] * len(fields),
            "COL_NAME": [f["col_name"] for f in fields],
            "DATA_TYPE": [f["type"] for f in fields],
        }
    )


def compare_and_unnest_parquet_schema(source_path, target_path):
    """Compare schemas of two Parquet files without loading full dataset."""
    return (
        _schema_to_df(pq.read_schema(source_path)),
        _schema_to_df(pq.read_schema(target_path)),
    )


def read_files(source_path, target_path):
    """Read source and target files using Polars."""
    if source_path.endswith(".parquet") and target_path.endswith(".parquet"):
        return compare_and_unnest_parquet_schema(source_path, target_path)

    source_df = pl.read_csv(source_path)
    target_df = pl.read_csv(target_path)

    for df_ref, name in [(source_df, "source"), (target_df, "target")]:
        if "DATA_TYPE" not in df_ref.columns:
            df_ref = df_ref.with_columns(pl.lit("N/A").alias("DATA_TYPE"))
        if name == "source":
            source_df = df_ref
        else:
            target_df = df_ref

    for df_ref, name in [(source_df, "source"), (target_df, "target")]:
        if "SCH" in df_ref.columns:
            df_ref = df_ref.with_columns(
                pl.concat_str([pl.col("SCH"), pl.lit("."), pl.col("TABLE_NAME")]).alias("SCH_TABLE")
            )
        else:
            df_ref = df_ref.with_columns(pl.col("TABLE_NAME").alias("SCH_TABLE"))
        if name == "source":
            source_df = df_ref
        else:
            target_df = df_ref

    return source_df, target_df


# ---------------------------------------------------------------------------
# Comparison logic
# ---------------------------------------------------------------------------


def compare_tables(source_df, target_df):
    """Compare tables between source and target."""
    source_tables = set(source_df["SCH_TABLE"].unique())
    target_tables = set(target_df["SCH_TABLE"].unique())
    return {
        "common": sorted(source_tables & target_tables),
        "source_only": sorted(source_tables - target_tables),
        "target_only": sorted(target_tables - source_tables),
    }


def compare_columns(source_df, target_df, table_name, type_mappings=None):
    """Compare columns for a specific table."""
    source_cols = source_df.filter(pl.col("SCH_TABLE") == table_name).select(["COL_NAME", "DATA_TYPE"])
    target_cols = target_df.filter(pl.col("SCH_TABLE") == table_name).select(["COL_NAME", "DATA_TYPE"])

    src_set = set(source_cols["COL_NAME"].to_list())
    tgt_set = set(target_cols["COL_NAME"].to_list())
    common = src_set & tgt_set

    mismatches = []
    for col in common:
        st = source_cols.filter(pl.col("COL_NAME") == col)["DATA_TYPE"].item()
        tt = target_cols.filter(pl.col("COL_NAME") == col)["DATA_TYPE"].item()
        if not are_types_compatible(st, tt, type_mappings):
            mismatches.append({"column": col, "source_type": st, "target_type": tt})

    return {
        "common": sorted(common),
        "source_only": sorted(src_set - tgt_set),
        "target_only": sorted(tgt_set - src_set),
        "datatype_mismatches": mismatches,
    }


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------


def _format_worksheet(ws):
    """Apply formatting to worksheet."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for column in ws.columns:
        column = list(column)
        max_len = max((len(str(c.value or "")) for c in column), default=0)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 2, 21.6)


def create_excel_report(comparison_results, source_df, target_df, file_name, type_mappings=None):
    """Create formatted Excel report."""
    wb = Workbook()

    # --- Table Comparison ---
    ws = wb.active
    ws.title = "Table Comparison"
    ws.append(["Category", "Table Name"])
    for cat in ("common", "source_only", "target_only"):
        label = cat.replace("_", " ").title()
        for t in comparison_results["tables"][cat]:
            ws.append([label, t])
    _format_worksheet(ws)

    # --- Column Comparison ---
    ws2 = wb.create_sheet("Column Comparison")
    ws2.append(["Table Name", "Column Name", "Status", "Source Type", "Target Type"])
    for tbl in comparison_results["columns"]:
        tn = tbl["table_name"]
        src = dict(
            zip(
                source_df.filter(pl.col("SCH_TABLE") == tn)["COL_NAME"],
                source_df.filter(pl.col("SCH_TABLE") == tn)["DATA_TYPE"],
            )
        )
        tgt = dict(
            zip(
                target_df.filter(pl.col("SCH_TABLE") == tn)["COL_NAME"],
                target_df.filter(pl.col("SCH_TABLE") == tn)["DATA_TYPE"],
            )
        )
        for col in sorted(set(list(src) + list(tgt))):
            st, tt = src.get(col, "N/A"), tgt.get(col, "N/A")
            if col in tbl["source_only"]:
                status = "Source Only"
            elif col in tbl["target_only"]:
                status = "Target Only"
            elif are_types_compatible(st, tt, type_mappings):
                status = "Matching"
            else:
                status = "Different Types"
            ws2.append([tn, col, status, st, tt])
    _format_worksheet(ws2)

    # --- Datatype Mismatches ---
    ws3 = wb.create_sheet("Datatype Mismatches")
    ws3.append(["Table Name", "Column Name", "Source Type", "Target Type"])
    for tbl in comparison_results["columns"]:
        for m in tbl["datatype_mismatches"]:
            ws3.append([tbl["table_name"], m["column"], m["source_type"], m["target_type"]])
    _format_worksheet(ws3)

    os.makedirs("results", exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = file_name.split("/")[-1] if "/" in file_name else file_name
    wb.save(f"results/{safe_name}_{ts}.xlsx")


# ---------------------------------------------------------------------------
# Core comparison runners
# ---------------------------------------------------------------------------


def _run_comparison(source_df, target_df, report_name, type_mappings=None):
    """Run table/column comparison and generate Excel report."""
    table_comparison = compare_tables(source_df, target_df)
    column_comparisons = []
    for tbl in table_comparison["common"]:
        cc = compare_columns(source_df, target_df, tbl, type_mappings)
        column_comparisons.append({"table_name": tbl, **cc})

    results = {"tables": table_comparison, "columns": column_comparisons}
    create_excel_report(results, source_df, target_df, report_name, type_mappings)
    return results


def colcompare_from_db(source_config_path, target_config_path, type_mappings=None):
    """Compare column schemas by fetching metadata directly from databases."""
    with Timer("Database column comparison"):
        source_config = load_config(source_config_path)
        target_config = load_config(target_config_path)

        tables_file = source_config.get("tables_file") or target_config.get("tables_file")
        if not tables_file:
            logger.error("tables_file must be specified in at least one config")
            return

        max_workers = source_config.get("max_workers", 4)
        df, source_tables, target_tables = _read_table_lists(tables_file)

        if target_tables is None:
            target_tables = source_tables

        source_df = fetch_metadata_parallel(source_config, source_tables, "source:", max_workers)
        target_df = fetch_metadata_parallel(target_config, target_tables, "target:", max_workers)

        report_name = Path(tables_file).stem
        _run_comparison(source_df, target_df, report_name, type_mappings)
        logger.info("Database column comparison complete")


def colcompare(args=None):
    """Run column comparison from files or database connections."""
    if isinstance(args, (list, type(None))):
        parser = argparse.ArgumentParser(
            description="Compare column schemas between CSV/Parquet files or databases",
            formatter_class=argparse.RawDescriptionHelpFormatter,
            epilog="""
Generates an Excel report with three sheets:
- Table Comparison: Lists tables present in source/target
- Column Comparison: Details of column presence and type matching
- Datatype Mismatches: Highlights columns with incompatible types

The report is saved to ./results/ with a timestamp in the filename.

To generate a default configuration file:
  dbqt colcompare --generate-config [--output PATH]
            """,
        )
        parser.add_argument("--generate-col-mappings", action="store_true",
                            help="Generate a default column type mappings configuration file")
        parser.add_argument("--output", "-o", default="colcompare_config.yaml",
                            help="Output path for config file (used with --generate-config)")
        parser.add_argument("source", nargs="?", help="Path to the source CSV/Parquet file")
        parser.add_argument("target", nargs="?", help="Path to the target CSV/Parquet file")
        parser.add_argument("--source-config", help="YAML config for source database")
        parser.add_argument("--target-config", help="YAML config for target database")
        parser.add_argument("--config", "-c", help="Path to type mappings configuration file")
        parser.add_argument("--verbose", "-v", action="store_true", help="Verbose logging")

        args = parser.parse_args(args)
        setup_logging(args.verbose)

        if args.generate_col_mappings:
            generate_config_file(args.output)
            return

        if args.source_config and args.target_config:
            tm = load_type_mappings(getattr(args, "config", None))
            colcompare_from_db(args.source_config, args.target_config, tm)
            return

        if not args.source or not args.target:
            parser.error(
                "Provide source and target files, or --source-config and --target-config"
            )

    type_mappings = load_type_mappings(getattr(args, "config", None))

    with Timer("Column comparison"):
        source_df, target_df = read_files(args.source, args.target)
        target_file_name = args.target.split(".")[0]
        _run_comparison(source_df, target_df, target_file_name, type_mappings)


def main(args=None):
    """Entry point for the colcompare tool."""
    colcompare(args)


if __name__ == "__main__":
    main()
